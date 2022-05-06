from communicationSystem.views import list_group
from django.shortcuts import render, redirect
from django.contrib import messages
from . import models as md
import openpyxl
import json


def view_marks(request, slug):
    """
        Queries results for different semesters for the given user.

        If, results are not available -
        1. Staff members are directed to a blank table with an error message.
        2. Students are redirected to their home page.
    """

    groups = list_group(request.user)

    username = "0"+str(slug) if len(str(slug)) == 10 else "00"+str(slug)
    student = md.Profile.objects.filter(user__username=username)

    params = {"student": student, "groups": groups}

    sem1 = md.Sem1.objects.filter(roll_no=slug)

    if len(sem1) == 0:
        messages.error(request, "Result Not Available")
        if request.user.is_superuser:
            return render(request, 'myApp/marks.html', params)
        else:
            return redirect("HomeStudent")

    sem2 = md.Sem2.objects.filter(roll_no=slug)
    sem3 = md.Sem3.objects.filter(roll_no=slug)
    sem4 = md.Sem4.objects.filter(roll_no=slug)
    sem5 = md.Sem5.objects.filter(roll_no=slug)
    sem6 = md.Sem6.objects.filter(roll_no=slug)
    sem7 = md.Sem7.objects.filter(roll_no=slug)
    sem8 = md.Sem8.objects.filter(roll_no=slug)

    params["marks"] = {"Semester-1": sem1, "Semester-2": sem2, "Semester-3": sem3, "Semester-4": sem4,
                       "Semester-5": sem5, "Semester-6": sem6, "Semester-7": sem7, "Semester-8": sem8}

    return render(request, 'myApp/marks.html', params)


def upload_marks(request):
    """
        Takes results in excel file and populates the data tables with the marks of respective students.
    """

    groups = list_group(request.user)
    params = {"groups": groups}

    if request.method == "POST":
        sem = request.POST.get("sem")
        file = request.FILES['excelFile']
        mid_term = request.POST.get("midTerm", "off")
        end_term = request.POST.get("endTerm", "off")
        practicals = request.POST.get("practicals", "off")

        # Getting data from excel
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Formatting data
        dict_data = format_data(sheet)
        dict_data = dict(sorted(dict_data.items(), key=sorter))

        # Preparing data to save
        row_max = sheet.max_row
        for i in range(0, row_max-1):
            roll = None
            marks_dict = {}

            for key in dict_data.keys():
                if key.lower() == "roll no":
                    roll = dict_data[key][i]

                else:
                    marks_dict[key] = dict_data[key][i]
            
            marks_dict = json.dumps(marks_dict)
            save_data(marks_dict, sem, mid_term, end_term, practicals, roll)

        messages.success(request, "Marks Uploaded Successfully")
        return redirect("UploadMarks")

    return render(request, 'myApp/upload.html', params)


# Utility functions

def sorter(item):
    return item[0]


def format_data(sheet):
    """
        1. Retrieves column labels from excel.
        2. Creates a dictionary with labels as keys and the data under the labels is taken as values.

        Example -
        Roll no.      Subject1    Subject2              dict = { "Roll No." : [01711502819, 02511502819],
        01711502819   25          23            ==>              "Subject1" : [25, 23]
        02522502819   23          24                             "Subject2" : [23, 24] }
    """

    col_max = sheet.max_column
    row_max = sheet.max_row

    dict_data = {}
    for i in range(1, col_max + 1):
        dict_data[f"{sheet.cell(1, i).value}"] = []

    j = 1
    for key in dict_data.keys():
        for i in range(2, row_max + 1):
            dict_data[key].append(sheet.cell(i, j).value)

        j += 1
        if j == col_max + 1:
            break

    return dict_data


def save_data(marks_dict, sem, mid_term, end_sem, practicals, roll):
    """
        Saves results for different semesters in the respective data tables.
    """
    if sem == "Sem1":
        if mid_term == "on":
            obj = md.Sem1(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem1.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem1.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()

    elif sem == "Sem2":
        if mid_term == "on":
            obj = md.Sem2(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem2.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem2.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()

    elif sem == "Sem3":
        if mid_term == "on":
            obj = md.Sem3(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem3.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem3.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()

    elif sem == "Sem4":
        if mid_term == "on":
            obj = md.Sem4(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem4.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem4.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()

    elif sem == "Sem5":
        if mid_term == "on":
            obj = md.Sem5(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem5.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem5.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()

    elif sem == "Sem6":
        if mid_term == "on":
            obj = md.Sem6(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem6.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem6.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()
    
    elif sem == "Sem7":
        if mid_term == "on":
            obj = md.Sem7(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem7.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem7.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()

    elif sem == "Sem8":
        if mid_term == "on":
            obj = md.Sem8(roll_no=roll, mid_sem=marks_dict)
            obj.save()

        elif end_sem == "on":
            obj = md.Sem8.objects.get(roll_no=roll)
            obj.end_sem = marks_dict
            obj.save()

        elif practicals == "on":
            obj = md.Sem8.objects.get(roll_no=roll)
            obj.practicals = marks_dict
            obj.save()
